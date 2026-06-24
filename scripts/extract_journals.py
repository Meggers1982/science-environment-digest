#!/usr/bin/env python3
"""
Extract journal data from PubMed_Journals_Categorized.xlsx for the five
Science & Environment Digest categories. Writes one CSV per sheet into data/.
"""

import csv
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl not installed — run: pip install openpyxl")

XLSX_PATH = Path("/Users/meaganmorris/Downloads/PubMed_Journals_Categorized.xlsx")
DATA_DIR  = Path(__file__).parent.parent / "data"

SHEETS = [
    "Environmental Health",
    "Public Health & Epidemiology",
    "Genetics & Molecular Biology",
    "Infectious Disease",
    "Pharmacology & Drug Therapy",
]

COLUMNS = ["Journal Title", "ISSN (Print)", "ISSN (Online)", "Categories"]


def safe_filename(name: str) -> str:
    return name.replace("/", "&").replace("\\", "&")


def extract_sheet(ws, sheet_name: str):
    headers = [cell.value for cell in ws[1]]
    col_idx = {}
    for col in COLUMNS:
        try:
            col_idx[col] = headers.index(col)
        except ValueError:
            raise ValueError(f"Column '{col}' not found in sheet '{sheet_name}'. Headers: {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append({col: (row[col_idx[col]] or "") for col in COLUMNS})
    return rows


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX_PATH)

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet_name}' not found — skipping")
            continue

        ws = wb[sheet_name]
        rows = extract_sheet(ws, sheet_name)

        out_path = DATA_DIR / f"{safe_filename(sheet_name)}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS)
            writer.writeheader()
            writer.writerows(rows)

        print(f"  {sheet_name}: {len(rows)} rows → {out_path.name}")

    print("Done.")


if __name__ == "__main__":
    main()
